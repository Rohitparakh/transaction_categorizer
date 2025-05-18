import streamlit as st
import pandas as pd
from classifier import classify_transactions
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import tempfile
import json
import uuid

from streamlit_local_storage import LocalStorage

st.set_page_config(page_title="Transaction Classifier", layout="centered")
st.title("💼 Transaction Classifier – Business Expense Categorizer")

# --- LocalStorage for category hierarchy ---
localS = LocalStorage()
categories_json = localS.getItem("categories")
if categories_json:
    categories = json.loads(categories_json)
else:
    categories = {}  # No default categories

uploaded_file = st.file_uploader("Upload your bank statement (.xlsx)", type=["xlsx"])

if uploaded_file:
    st.subheader("Business Categories & Keywords (Category > Subcategory > Keywords)")
    st.write("Add/edit categories and keywords. These will be saved in your browser.")

    if "category_inputs" not in st.session_state:
        st.session_state.category_inputs = categories.copy()

    def save_categories():
        localS.setItem("categories", json.dumps(st.session_state.category_inputs))
        st.toast("Saved to localStorage!") 

    # Helper function to init session state for a given key with default value
    def init_state(key, default_val):
        if key not in st.session_state:
            st.session_state[key] = default_val

    # --- CATEGORY/SUBCATEGORY EDITOR WITH COLLAPSIBLE CATEGORIES ---
    for cat in list(st.session_state.category_inputs.keys()):
        with st.expander(f"🗂️ {cat}", expanded=False):
            # Rename category
            rename_cat_key = f"rename_cat_{cat}"
            init_state(rename_cat_key, cat)
            new_cat_name = st.text_input("Rename category", value=st.session_state[rename_cat_key], key=rename_cat_key)
            
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button("Rename Category", key=f"btn_rename_{cat}"):
                    if new_cat_name and new_cat_name != cat and new_cat_name not in st.session_state.category_inputs:
                        st.session_state.category_inputs[new_cat_name] = st.session_state.category_inputs.pop(cat)
                        # Move over keyword input fields in session_state
                        for subcat in list(st.session_state.category_inputs[new_cat_name].keys()):
                            old_key = f"{cat}_{subcat}_keywords"
                            new_key = f"{new_cat_name}_{subcat}_keywords"
                            if old_key in st.session_state:
                                st.session_state[new_key] = st.session_state.pop(old_key)
                        # Clear old rename input key to avoid conflicts
                        del st.session_state[rename_cat_key]
                        save_categories()
                        st.rerun()
            with col2:
                if st.button("Delete Category", key=f"delete_cat_{cat}"):
                    del st.session_state.category_inputs[cat]
                    # Also clear any related keys from session_state
                    # Clear rename key if exists
                    if rename_cat_key in st.session_state:
                        del st.session_state[rename_cat_key]
                    save_categories()
                    st.rerun()

            # --- Subcategories (NO expanders here) ---
            for subcat in list(st.session_state.category_inputs[cat].keys()):
                st.markdown(f"&nbsp;&nbsp;&nbsp;**• _{subcat}_**", unsafe_allow_html=True)

                sub_rename_key = f"rename_subcat_{cat}_{subcat}"
                init_state(sub_rename_key, subcat)
                sub_col1, sub_col2 = st.columns([1, 1])
                with sub_col1:
                    new_subcat_name = st.text_input(
                        f"Rename subcategory '{subcat}' under {cat}",
                        value=st.session_state[sub_rename_key],
                        key=sub_rename_key
                    )
                    if st.button("Rename Subcategory", key=f"btn_rename_subcat_{cat}_{subcat}"):
                        if new_subcat_name and new_subcat_name != subcat and new_subcat_name not in st.session_state.category_inputs[cat]:
                            st.session_state.category_inputs[cat][new_subcat_name] = st.session_state.category_inputs[cat].pop(subcat)
                            old_key = f"{cat}_{subcat}_keywords"
                            new_key = f"{cat}_{new_subcat_name}_keywords"
                            if old_key in st.session_state:
                                st.session_state[new_key] = st.session_state.pop(old_key)
                            # Clear old rename input key to avoid conflicts
                            if sub_rename_key in st.session_state:
                                del st.session_state[sub_rename_key]
                            save_categories()
                            st.rerun()
                with sub_col2:
                    if st.button("Delete Subcategory", key=f"del_{cat}_{subcat}"):
                        del st.session_state.category_inputs[cat][subcat]
                        # Clear rename key if exists
                        if sub_rename_key in st.session_state:
                            del st.session_state[sub_rename_key]
                        save_categories()
                        st.rerun()

                # Keywords editor
                keywords_key = f"{cat}_{subcat}_keywords"
                default_keywords = ", ".join(st.session_state.category_inputs[cat][subcat])
                init_state(keywords_key, default_keywords)
                updated_keywords = st.text_area(
                    "",
                    value=st.session_state[keywords_key],
                    key=keywords_key,
                    label_visibility="collapsed",
                    height=68
                )
                # Update session_state and category inputs on change
                if updated_keywords != st.session_state[keywords_key]:
                    st.session_state[keywords_key] = updated_keywords
                # Update category_inputs list
                st.session_state.category_inputs[cat][subcat] = [k.strip() for k in updated_keywords.split(",") if k.strip()]

            # Add new subcategory input
            add_subcat_key = f"add_subcat_{cat}"
            init_state(add_subcat_key, "")
            st.markdown("&nbsp;&nbsp;&nbsp;_Add subcategory:_", unsafe_allow_html=True)
            new_subcat = st.text_input("", key=add_subcat_key, label_visibility="collapsed", placeholder="Type subcategory name")
            if st.button(f"Add Subcategory to {cat}", key=f"btn_add_subcat_{cat}"):
                if new_subcat and new_subcat not in st.session_state.category_inputs[cat]:
                    st.session_state.category_inputs[cat][new_subcat] = []
                    # Clear input field after adding
                    # st.session_state[add_subcat_key] = ""
                    save_categories()
                    st.rerun()

            st.markdown("---")

    # Add new category input
    add_cat_key = "new_cat"
    init_state(add_cat_key, "")
    st.markdown("#### ➕ Add new Expense Category")
    new_cat = st.text_input("", key=add_cat_key, label_visibility="collapsed", placeholder="Type category name")
    if st.button("Add Category"):
        if new_cat and new_cat not in st.session_state.category_inputs:
            st.session_state.category_inputs[new_cat] = {}
            # Clear input field after adding
            # st.session_state[add_cat_key] = ""
            save_categories()
            st.rerun()

    if st.button("Save Categories"):
        save_categories()
        st.success("Categories saved to your browser!")

    categories_for_classifier = st.session_state.category_inputs

    header_row = st.number_input("Which row contains headers?", min_value=1, step=1, value=17)

    try:
        # Save uploaded file to temp
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getbuffer())
            tmp_path = tmp.name

        df_raw = pd.read_excel(tmp_path, header=None, dtype=str)
        pre_header = df_raw.iloc[:header_row-1, :]
        df = pd.read_excel(tmp_path, header=header_row - 1, dtype=str)
        st.success("Headers loaded successfully. Map columns below.")

        col_options = df.columns.tolist()
        col_serial = st.selectbox("🔢 Serial Number column", col_options, index=col_options.index("S.N.") if "S.N." in col_options else 0)
        col_remarks = st.selectbox("📝 Transaction Remarks column", col_options, index=col_options.index("Transaction Remarks") if "Transaction Remarks" in col_options else 0)
        col_withdrawal = st.selectbox("💸 Withdrawal Amount column", col_options, index=col_options.index("Withdrawal Amt (INR)") if "Withdrawal Amt (INR)" in col_options else 0)
        col_deposit = st.selectbox("💰 Deposit Amount column", col_options, index=col_options.index("Deposit Amt (INR)") if "Deposit Amt (INR)" in col_options else 0)

        if st.button("🔍 Process Transactions"):
            processed_df = classify_transactions(
                df.copy(),
                col_remarks,
                col_withdrawal,
                col_deposit,
                col_serial,
                categories_for_classifier
            )
            st.success("✅ File processed successfully!")
            st.dataframe(processed_df)

            wb = load_workbook(tmp_path)
            ws = wb.active

            start_col = ws.max_column + 1
            header_row_idx = header_row

            # headers = ['Expense Type', 'Expense Category', 'Expense Subcategory','Keyword', 'Match Score']
            headers = ['Expense Type', 'Expense Category', 'Expense Subcategory', 'Remarks']           
            # Apply same style as last original column
            style_source_col = start_col - 1

            for i, header in enumerate(headers):
                new_col = start_col + i
                header_cell = ws.cell(row=header_row_idx, column=new_col)
                header_cell.value = header

                # Copy style from last column's header
                source_header_cell = ws.cell(row=header_row_idx, column=style_source_col)
                header_cell.font = copy(source_header_cell.font)
                header_cell.border = copy(source_header_cell.border)
                header_cell.fill = copy(source_header_cell.fill)
                header_cell.alignment = copy(source_header_cell.alignment)

                for row_idx, row in processed_df.iterrows():
                    cell = ws.cell(row=header_row_idx + 1 + row_idx, column=new_col)
                    cell.value = row[headers[i]]

                    # Copy style from adjacent column
                    source_cell = ws.cell(row=header_row_idx + 1 + row_idx, column=style_source_col)
                    cell.font = copy(source_cell.font)
                    cell.border = copy(source_cell.border)
                    cell.fill = copy(source_cell.fill)
                    cell.alignment = copy(source_cell.alignment)
                max_length = len(str(header))
                for row_idx in range(header_row_idx + 1, header_row_idx + 1 + len(processed_df)):
                    cell_value = ws.cell(row=row_idx, column=new_col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                adjusted_width = (max_length + 2)  # Add some padding
                col_letter = get_column_letter(new_col)
                ws.column_dimensions[col_letter].width = adjusted_width                
            output_stream = BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)

            st.download_button(
                label="📥 Download Processed File",
                data=output_stream,
                file_name="processed_transactions.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        if uploaded_file:
            st.error(f"Error processing file: {e}")

